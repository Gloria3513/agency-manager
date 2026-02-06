"""
엑셀 내보내기
데이터를 Excel 형식으로 내보내기
"""

from datetime import datetime
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path


class ExcelExporter:
    """엑셀 내보내기"""

    def __init__(self):
        self.wb = None
        self.ws = None

    def create_workbook(self):
        """새 워크북 생성"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"

    def add_title(self, title: str, row: int = 1):
        """제목 추가"""
        self.ws.merge_cells(f"A{row}:D{row}")
        cell = self.ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def add_section_title(self, title: str, row: int):
        """섹션 제목 추가"""
        cell = self.ws.cell(row=row, column=1)
        cell.value = title
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    def add_headers(self, headers: List[str], row: int):
        """헤더 추가"""
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def add_data_row(self, data: List, row: int):
        """데이터 행 추가"""
        for col, value in enumerate(data, start=1):
            cell = self.ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def auto_fit_columns(self):
        """열 너비 자동 조정"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def save(self, filepath: str):
        """파일 저장"""
        self.wb.save(filepath)


def export_clients_to_excel(clients: List[Dict], filepath: str = None) -> str:
    """고객 목록 엑셀 내보내기"""
    if filepath is None:
        filepath = f"clients_{datetime.now().strftime('%Y%m%d')}.xlsx"

    exporter = ExcelExporter()
    exporter.create_workbook()

    # 제목
    exporter.add_title("고객 목록", 1)

    # 헤더
    headers = ["ID", "이름", "이메일", "연락처", "회사", "상태", "등록일"]
    exporter.add_headers(headers, 3)

    # 데이터
    for idx, client in enumerate(clients, start=4):
        exporter.add_data_row([
            client.get('id'),
            client.get('name'),
            client.get('email'),
            client.get('phone'),
            client.get('company'),
            client.get('status'),
            client.get('created_at', '')[:10]
        ], idx)

    exporter.auto_fit_columns()
    exporter.save(filepath)

    return filepath


def export_projects_to_excel(projects: List[Dict], filepath: str = None) -> str:
    """프로젝트 목록 엑셀 내보내기"""
    if filepath is None:
        filepath = f"projects_{datetime.now().strftime('%Y%m%d')}.xlsx"

    exporter = ExcelExporter()
    exporter.create_workbook()

    # 제목
    exporter.add_title("프로젝트 목록", 1)

    # 헤더
    headers = ["ID", "프로젝트명", "고객", "상태", "진행률", "계약 금액", "시작일", "종료일"]
    exporter.add_headers(headers, 3)

    # 데이터
    for idx, project in enumerate(projects, start=4):
        exporter.add_data_row([
            project.get('id'),
            project.get('name'),
            project.get('client_name'),
            project.get('status'),
            f"{project.get('progress', 0)}%",
            f"{project.get('total_contract_amount', 0):,.0f}원",
            project.get('start_date', ''),
            project.get('end_date', '')
        ], idx)

    exporter.auto_fit_columns()
    exporter.save(filepath)

    return filepath


def export_payments_to_excel(payments: List[Dict], filepath: str = None) -> str:
    """결제 내역 엑셀 내보내기"""
    if filepath is None:
        filepath = f"payments_{datetime.now().strftime('%Y%m%d')}.xlsx"

    exporter = ExcelExporter()
    exporter.create_workbook()

    # 제목
    exporter.add_title("결제 내역", 1)

    # 헤더
    headers = ["송장번호", "프로젝트", "고객", "유형", "금액", "입금 예정일", "상태", "입금일"]
    exporter.add_headers(headers, 3)

    # 데이터
    for idx, payment in enumerate(payments, start=4):
        exporter.add_data_row([
            payment.get('invoice_number'),
            payment.get('project_name'),
            payment.get('client_name'),
            payment.get('payment_type'),
            f"{payment.get('amount', 0):,.0f}원",
            payment.get('due_date', ''),
            payment.get('status'),
            payment.get('paid_date', '')
        ], idx)

    # 합계 행
    total_row = len(payments) + 4
    total_amount = sum(p.get('amount', 0) for p in payments)
    exporter.ws.merge_cells(f"A{total_row}:G{total_row}")
    cell = exporter.ws.cell(row=total_row, column=1)
    cell.value = f"총액: {total_amount:,.0f}원"
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    exporter.auto_fit_columns()
    exporter.save(filepath)

    return filepath


def export_time_entries_to_excel(time_entries: List[Dict], filepath: str = None) -> str:
    """시간 기록 엑셀 내보내기"""
    if filepath is None:
        filepath = f"time_entries_{datetime.now().strftime('%Y%m%d')}.xlsx"

    exporter = ExcelExporter()
    exporter.create_workbook()

    # 제목
    exporter.add_title("시간 기록", 1)

    # 헤더
    headers = ["날짜", "프로젝트", "작업 내용", "소요 시간(분)", "시간", "청구 가능 여부"]
    exporter.add_headers(headers, 3)

    total_minutes = 0

    # 데이터
    for idx, entry in enumerate(time_entries, start=4):
        minutes = entry.get('duration_minutes', 0)
        hours = minutes / 60
        total_minutes += minutes

        exporter.add_data_row([
            entry.get('entry_date', '')[:10],
            entry.get('project_name', ''),
            entry.get('title'),
            minutes,
            f"{hours:.1f}",
            "예" if entry.get('billable') else "아니오"
        ], idx)

    # 합계 행
    total_row = len(time_entries) + 4
    total_hours = total_minutes / 60
    exporter.ws.merge_cells(f"A{total_row}:F{total_row}")
    cell = exporter.ws.cell(row=total_row, column=1)
    cell.value = f"총 시간: {total_hours:.1f}시간 ({total_minutes}분)"
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    exporter.auto_fit_columns()
    exporter.save(filepath)

    return filepath
